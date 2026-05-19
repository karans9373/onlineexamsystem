from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any
import json

from openai import OpenAI
from openpyxl import load_workbook
from pypdf import PdfReader
from docx import Document
from sqlalchemy.orm import Session

from .config import get_settings
from .models import Exam, ExamAttempt, Notification, Question, SuspiciousActivityLog

settings = get_settings()


def serialize_question(question: Question, include_answer: bool = False) -> dict[str, Any]:
    payload = {
        "id": question.id,
        "order_index": question.order_index,
        "text_en": question.text_en,
        "text_hi": question.text_hi,
        "image_url": question.image_url,
        "options": question.options,
        "difficulty": question.difficulty,
        "topic": question.topic,
        "marks": question.marks,
        "explanation": question.explanation,
    }
    if include_answer:
        payload["correct_option"] = question.correct_option
    return payload


def serialize_exam(exam: Exam, include_questions: bool = False, include_answers: bool = False) -> dict[str, Any]:
    payload = {
        "id": exam.id,
        "title": exam.title,
        "description": exam.description,
        "subject": exam.subject,
        "exam_type": exam.exam_type,
        "status": exam.status,
        "duration_minutes": exam.duration_minutes,
        "total_questions": exam.total_questions,
        "total_marks": exam.total_marks,
        "negative_marking": exam.negative_marking,
        "language_mode": exam.language_mode,
        "instructions": exam.instructions,
        "start_at": exam.start_at.isoformat() if exam.start_at else None,
        "end_at": exam.end_at.isoformat() if exam.end_at else None,
        "created_by": exam.created_by,
        "creator_name": exam.creator.full_name if exam.creator else None,
        "is_published": exam.is_published,
    }
    if include_questions:
        payload["questions"] = [serialize_question(question, include_answer=include_answers) for question in exam.questions]
    return payload


def create_notification(db: Session, title: str, message: str, type_: str = "info", user_id: int | None = None):
    notification = Notification(user_id=user_id, title=title, message=message, type=type_)
    db.add(notification)
    db.flush()
    return notification


def log_suspicious_activity(
    db: Session,
    attempt: ExamAttempt,
    activity_type: str,
    detail: str,
    severity: str = "medium",
):
    record = SuspiciousActivityLog(
        attempt_id=attempt.id,
        exam_id=attempt.exam_id,
        student_id=attempt.student_id,
        activity_type=activity_type,
        detail=detail,
        severity=severity,
    )
    db.add(record)
    create_notification(
        db,
        title="Suspicious activity detected",
        message=f"{attempt.student.full_name} triggered {activity_type}: {detail}",
        type_="warning",
        user_id=attempt.exam.created_by,
    )
    return record


def build_attempt_analysis(attempt: ExamAttempt) -> dict[str, Any]:
    questions = attempt.exam.questions
    answers = attempt.answers or {}
    time_spent = attempt.time_spent or {}
    attempted = sum(1 for question in questions if answers.get(str(question.id)))
    correct = 0
    wrong = 0
    subject_breakdown: dict[str, dict[str, float]] = {}
    review = []

    for question in questions:
        selected = answers.get(str(question.id))
        is_correct = selected == question.correct_option
        if selected:
            if is_correct:
                correct += 1
            else:
                wrong += 1
        topic_key = question.topic or "General"
        bucket = subject_breakdown.setdefault(topic_key, {"correct": 0, "wrong": 0, "total": 0})
        bucket["total"] += 1
        if selected:
            if is_correct:
                bucket["correct"] += 1
            else:
                bucket["wrong"] += 1
        review.append(
            {
                "question_id": question.id,
                "question": question.text_en,
                "selected": selected,
                "correct": question.correct_option,
                "is_correct": is_correct,
                "time_spent_seconds": time_spent.get(str(question.id), 0),
                "topic": topic_key,
            }
        )

    marks = correct - (wrong * attempt.exam.negative_marking)
    percentage = round((max(marks, 0) / max(attempt.exam.total_marks, 1)) * 100, 2)
    accuracy = round((correct / max(attempted, 1)) * 100, 2) if attempted else 0
    weakest_topics = sorted(subject_breakdown.items(), key=lambda item: item[1]["correct"] / max(item[1]["total"], 1))[:3]
    suggestions = [
        f"Revise {topic} with timed sets and focus on eliminating distractors."
        for topic, _ in weakest_topics
    ] or ["Maintain current pace and attempt one full mock every alternate day."]

    return {
        "score": round(max(marks, 0), 2),
        "attempted": attempted,
        "correct": correct,
        "wrong": wrong,
        "percentage": percentage,
        "accuracy": accuracy,
        "subject_breakdown": subject_breakdown,
        "weak_topics": [topic for topic, _ in weakest_topics],
        "suggestions": suggestions,
        "review": review,
    }


def refresh_exam_rankings(db: Session, exam: Exam):
    submitted_attempts = (
        db.query(ExamAttempt)
        .filter(ExamAttempt.exam_id == exam.id, ExamAttempt.status == "submitted")
        .order_by(ExamAttempt.score.desc(), ExamAttempt.submitted_at.asc())
        .all()
    )
    total = len(submitted_attempts)
    for index, attempt in enumerate(submitted_attempts, start=1):
        attempt.rank = index
        attempt.percentile = round(((total - index) / max(total, 1)) * 100, 2)
    db.flush()


def parse_uploaded_questions(filename: str, file_bytes: bytes) -> list[dict[str, Any]]:
    name = filename.lower()
    if name.endswith(".pdf"):
        text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(file_bytes)).pages)
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        parsed = []
        current = []
        for line in lines:
            current.append(line)
            if len(current) == 5:
                parsed.append(
                    {
                        "text_en": current[0],
                        "text_hi": current[0],
                        "options": [{"key": key, "label": value} for key, value in zip(["A", "B", "C", "D"], current[1:5])],
                        "correct_option": "A",
                        "difficulty": "medium",
                        "topic": "Imported",
                        "marks": 1,
                    }
                )
                current = []
        return parsed

    if name.endswith(".docx"):
        document = Document(BytesIO(file_bytes))
        text = "\n".join(paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip())
        file_bytes = text.encode("utf-8")

    if name.endswith(".doc"):
        # Best-effort legacy Word fallback: recover visible text when possible.
        text = file_bytes.decode("latin-1", errors="ignore")
        text = text.replace("\x00", " ")
        file_bytes = text.encode("utf-8")

    if name.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(file_bytes))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        header_map = {header.lower(): index for index, header in enumerate(headers)}
        parsed = []
        for row in rows[1:]:
            parsed.append(
                {
                    "text_en": row[header_map["question_en"]],
                    "text_hi": row[header_map.get("question_hi", header_map["question_en"])],
                    "options": [
                        {"key": "A", "label": row[header_map["option_a"]]},
                        {"key": "B", "label": row[header_map["option_b"]]},
                        {"key": "C", "label": row[header_map["option_c"]]},
                        {"key": "D", "label": row[header_map["option_d"]]},
                    ],
                    "correct_option": str(row[header_map["correct_option"]]).strip().upper(),
                    "difficulty": str(row[header_map.get("difficulty", 0)] or "medium"),
                    "topic": str(row[header_map.get("topic", 0)] or "Imported"),
                    "marks": float(row[header_map.get("marks", 0)] or 1),
                    "explanation": str(row[header_map.get("explanation", 0)] or ""),
                }
            )
        return parsed

    text = file_bytes.decode("utf-8", errors="ignore")
    parsed = []
    for block in [section.strip() for section in text.split("\n\n") if section.strip()]:
        lines = [line.strip() for line in block.splitlines() if line.strip()]
        if len(lines) < 5:
            continue
        parsed.append(
            {
                "text_en": lines[0],
                "text_hi": lines[0],
                "options": [{"key": key, "label": label} for key, label in zip(["A", "B", "C", "D"], lines[1:5])],
                "correct_option": lines[5].replace("Answer:", "").strip().upper() if len(lines) > 5 else "A",
                "difficulty": "medium",
                "topic": "Imported",
                "marks": 1,
            }
        )
    return parsed


def _extract_topics(prompt: str, subject: str) -> list[str]:
    text = f"{subject} {prompt}".lower()
    known_topics = [
        "algebra",
        "mensuration",
        "geometry",
        "trigonometry",
        "probability",
        "statistics",
        "arithmetic",
        "percentage",
        "profit and loss",
        "time and distance",
        "time and work",
        "science",
        "physics",
        "chemistry",
        "biology",
        "english",
        "grammar",
        "hindi",
    ]
    matches = [topic for topic in known_topics if topic in text]
    return matches or [subject.lower()]


def _topic_bank() -> dict[str, list[dict[str, Any]]]:
    return {
        "algebra": [
            {
                "text_en": "If 3x + 5 = 20, what is the value of x?",
                "text_hi": "यदि 3x + 5 = 20 है, तो x का मान क्या है?",
                "options": [{"key": "A", "label": "3"}, {"key": "B", "label": "4"}, {"key": "C", "label": "5"}, {"key": "D", "label": "6"}],
                "correct_option": "C",
                "difficulty": "easy",
                "topic": "Algebra",
                "explanation": "Subtract 5 from both sides and divide by 3.",
            },
            {
                "text_en": "If x + y = 12 and x - y = 4, then x is:",
                "text_hi": "यदि x + y = 12 और x - y = 4 है, तो x है:",
                "options": [{"key": "A", "label": "6"}, {"key": "B", "label": "7"}, {"key": "C", "label": "8"}, {"key": "D", "label": "9"}],
                "correct_option": "C",
                "difficulty": "medium",
                "topic": "Algebra",
                "explanation": "Adding both equations gives 2x = 16, so x = 8.",
            },
            {
                "text_en": "If x^2 - 9 = 0, which value is a solution?",
                "text_hi": "यदि x^2 - 9 = 0 है, तो कौन सा मान एक हल है?",
                "options": [{"key": "A", "label": "2"}, {"key": "B", "label": "3"}, {"key": "C", "label": "4"}, {"key": "D", "label": "5"}],
                "correct_option": "B",
                "difficulty": "medium",
                "topic": "Algebra",
                "explanation": "x^2 = 9, so x = 3 or -3. From the options, 3 is correct.",
            },
        ],
        "mensuration": [
            {
                "text_en": "The area of a rectangle with length 12 cm and breadth 5 cm is:",
                "text_hi": "12 सेमी लंबाई और 5 सेमी चौड़ाई वाले आयत का क्षेत्रफल है:",
                "options": [{"key": "A", "label": "17 cm²"}, {"key": "B", "label": "60 cm²"}, {"key": "C", "label": "34 cm²"}, {"key": "D", "label": "70 cm²"}],
                "correct_option": "B",
                "difficulty": "easy",
                "topic": "Mensuration",
                "explanation": "Area of a rectangle is length multiplied by breadth.",
            },
            {
                "text_en": "The circumference of a circle of radius 7 cm is: (Use pi = 22/7)",
                "text_hi": "7 सेमी त्रिज्या वाले वृत्त की परिधि है: (pi = 22/7 लें)",
                "options": [{"key": "A", "label": "22 cm"}, {"key": "B", "label": "44 cm"}, {"key": "C", "label": "49 cm"}, {"key": "D", "label": "154 cm"}],
                "correct_option": "B",
                "difficulty": "medium",
                "topic": "Mensuration",
                "explanation": "Circumference = 2 × pi × r = 44 cm.",
            },
            {
                "text_en": "The volume of a cube with side 4 cm is:",
                "text_hi": "4 सेमी भुजा वाले घन का आयतन है:",
                "options": [{"key": "A", "label": "16 cm³"}, {"key": "B", "label": "32 cm³"}, {"key": "C", "label": "64 cm³"}, {"key": "D", "label": "48 cm³"}],
                "correct_option": "C",
                "difficulty": "easy",
                "topic": "Mensuration",
                "explanation": "Volume of a cube is side cubed, so 4 × 4 × 4 = 64.",
            },
        ],
        "percentage": [
            {
                "text_en": "What is 15% of 240?",
                "text_hi": "240 का 15% कितना है?",
                "options": [{"key": "A", "label": "24"}, {"key": "B", "label": "30"}, {"key": "C", "label": "36"}, {"key": "D", "label": "42"}],
                "correct_option": "C",
                "difficulty": "easy",
                "topic": "Percentage",
                "explanation": "15% of 240 equals 36.",
            },
            {
                "text_en": "A number increases from 80 to 100. The percentage increase is:",
                "text_hi": "एक संख्या 80 से 100 हो जाती है। प्रतिशत वृद्धि है:",
                "options": [{"key": "A", "label": "20%"}, {"key": "B", "label": "25%"}, {"key": "C", "label": "30%"}, {"key": "D", "label": "40%"}],
                "correct_option": "B",
                "difficulty": "medium",
                "topic": "Percentage",
                "explanation": "Increase = 20. Percentage increase = 20/80 × 100 = 25%.",
            },
        ],
        "time and distance": [
            {
                "text_en": "A car travels 150 km in 3 hours. Its speed is:",
                "text_hi": "एक कार 3 घंटे में 150 किमी चलती है। उसकी चाल है:",
                "options": [{"key": "A", "label": "40 km/h"}, {"key": "B", "label": "45 km/h"}, {"key": "C", "label": "50 km/h"}, {"key": "D", "label": "55 km/h"}],
                "correct_option": "C",
                "difficulty": "easy",
                "topic": "Time and Distance",
                "explanation": "Speed = distance divided by time = 50 km/h.",
            },
            {
                "text_en": "If a train covers 360 km in 4.5 hours, its average speed is:",
                "text_hi": "यदि एक ट्रेन 4.5 घंटे में 360 किमी तय करती है, तो उसकी औसत चाल है:",
                "options": [{"key": "A", "label": "70 km/h"}, {"key": "B", "label": "75 km/h"}, {"key": "C", "label": "80 km/h"}, {"key": "D", "label": "90 km/h"}],
                "correct_option": "C",
                "difficulty": "medium",
                "topic": "Time and Distance",
                "explanation": "360 divided by 4.5 equals 80 km/h.",
            },
        ],
        "science": [
            {
                "text_en": "Which gas is essential for human respiration?",
                "text_hi": "मानव श्वसन के लिए कौन सी गैस आवश्यक है?",
                "options": [{"key": "A", "label": "Nitrogen"}, {"key": "B", "label": "Oxygen"}, {"key": "C", "label": "Hydrogen"}, {"key": "D", "label": "Carbon monoxide"}],
                "correct_option": "B",
                "difficulty": "easy",
                "topic": "Science",
                "explanation": "Humans require oxygen for respiration.",
            },
            {
                "text_en": "The process by which green plants prepare food is called:",
                "text_hi": "वह प्रक्रिया जिससे हरे पौधे भोजन बनाते हैं, कहलाती है:",
                "options": [{"key": "A", "label": "Respiration"}, {"key": "B", "label": "Photosynthesis"}, {"key": "C", "label": "Transpiration"}, {"key": "D", "label": "Germination"}],
                "correct_option": "B",
                "difficulty": "easy",
                "topic": "Science",
                "explanation": "Plants use photosynthesis to prepare food.",
            },
        ],
        "english": [
            {
                "text_en": "Choose the correct synonym of 'rapid'.",
                "text_hi": "‘rapid’ का सही समानार्थी शब्द चुनिए।",
                "options": [{"key": "A", "label": "Slow"}, {"key": "B", "label": "Fast"}, {"key": "C", "label": "Weak"}, {"key": "D", "label": "Dull"}],
                "correct_option": "B",
                "difficulty": "easy",
                "topic": "English",
                "explanation": "Rapid means fast.",
            },
            {
                "text_en": "Choose the sentence with correct grammar.",
                "text_hi": "सही व्याकरण वाला वाक्य चुनिए।",
                "options": [{"key": "A", "label": "She go to school daily."}, {"key": "B", "label": "She goes to school daily."}, {"key": "C", "label": "She going to school daily."}, {"key": "D", "label": "She gone to school daily."}],
                "correct_option": "B",
                "difficulty": "medium",
                "topic": "English",
                "explanation": "With 'she', the correct verb form is 'goes'.",
            },
        ],
    }


def normalize_generated_questions(questions: list[dict[str, Any]], subject: str, bilingual: bool) -> list[dict[str, Any]]:
    normalized = []
    for index, question in enumerate(questions, start=1):
        options = question.get("options") or []
        if len(options) != 4:
            options = [
                {"key": "A", "label": "Option A"},
                {"key": "B", "label": "Option B"},
                {"key": "C", "label": "Option C"},
                {"key": "D", "label": "Option D"},
            ]
        normalized.append(
            {
                "text_en": question.get("text_en") or f"{subject} Question {index}",
                "text_hi": question.get("text_hi") if bilingual else None,
                "options": [
                    {"key": option.get("key", chr(65 + opt_index)), "label": option.get("label", f"Option {chr(65 + opt_index)}")}
                    for opt_index, option in enumerate(options[:4])
                ],
                "correct_option": (question.get("correct_option") or "A").upper(),
                "difficulty": question.get("difficulty") or "medium",
                "topic": question.get("topic") or subject,
                "marks": float(question.get("marks", 1)),
                "explanation": question.get("explanation") or "Teacher review explanation.",
            }
        )
    return normalized


def fallback_generate_questions(prompt: str, question_count: int, subject: str, bilingual: bool = True):
    topics = _extract_topics(prompt, subject)
    bank = _topic_bank()
    generated: list[dict[str, Any]] = []

    while len(generated) < question_count:
        current_topic = topics[len(generated) % len(topics)]
        pool = bank.get(current_topic) or bank.get(subject.lower()) or bank["science"]
        template = pool[len(generated) % len(pool)].copy()
        template["text_hi"] = template.get("text_hi") if bilingual else None
        template["marks"] = 1
        generated.append(template)

    return generated[:question_count]


def generate_questions_with_ai(prompt: str, question_count: int, subject: str, language_mode: str):
    bilingual = language_mode == "bilingual"
    if not settings.openai_api_key:
        return normalize_generated_questions(
            fallback_generate_questions(prompt, question_count, subject, bilingual=bilingual),
            subject,
            bilingual,
        )

    try:
        client = OpenAI(api_key=settings.openai_api_key)
        completion = client.responses.create(
            model=settings.openai_model,
            input=[
                {
                    "role": "system",
                    "content": (
                        "You generate high-quality multiple-choice questions as valid JSON only. "
                        "Return an object with a 'questions' array. Each question must include text_en, text_hi, "
                        "options as four items with keys A-D, correct_option, difficulty, topic, marks, explanation. "
                        "Questions must be academically correct, non-random, and suitable for classroom or CBT exam use."
                    ),
                },
                {
                    "role": "user",
                    "content": (
                        f"Subject: {subject}\n"
                        f"Count: {question_count}\n"
                        f"Language mode: {language_mode}\n"
                        f"Teacher request: {prompt}\n"
                        "Keep the paper balanced in difficulty and ensure every distractor is plausible."
                    ),
                },
            ],
        )
        text = completion.output_text
        payload = json.loads(text)
        return normalize_generated_questions(payload["questions"], subject, bilingual)
    except Exception:
        return normalize_generated_questions(
            fallback_generate_questions(prompt, question_count, subject, bilingual=bilingual),
            subject,
            bilingual,
        )
